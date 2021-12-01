import openpyxl
from openpyxl.utils import column_index_from_string

NAME_COL = 'E'
TITLE_COL = 'G'
USAGE_COL = 'H'
RESULT_FILE = 'leader_result.xlsx'

def extract_from_filter(sheet):
    keywords = set()
    names = list(sheet.columns)[column_index_from_string(NAME_COL)-1]
    titles = list(sheet.columns)[column_index_from_string(TITLE_COL)-1]
    keywords.update([name.value for name in names])
    keywords.update([title.value[-2:] for title in titles])
    keywords.add('副总')
    return keywords

def main():
    work_book = openpyxl.load_workbook('全省2021年统计表汇总.xlsx')
    data_sheets = [work_book['1-4'], work_book['5-8'], work_book['9-11']]
    filter_sheet = work_book['名单']

    leader_book = openpyxl.Workbook()
    leader_sheet = leader_book.active
    leader_sheet.append([cell.value for cell in list(data_sheets[0].rows)[0]])

    keywords = extract_from_filter(filter_sheet)

    for data in data_sheets:
        for row in data.iter_rows():
            usage = row[column_index_from_string(USAGE_COL)-1].value
            usage = str(usage)

            found = False
            for keyword in keywords:
                if keyword in usage:
                    found = True
                    break
    
            if found:
                leader_sheet.append([cell.value for cell in row])

    leader_book.save(RESULT_FILE)
    print('{} saved'.format(RESULT_FILE))
    

if __name__ == '__main__':
    main()

