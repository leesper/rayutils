import openpyxl
from openpyxl.utils import column_index_from_string

LOWER_BOUND = 10
UPPER_BOUND = 1000
AREA_FIELD = 'F'
ARCH_FIELD = 'G'
AGENT_FIELD = 'I'


def main():
    workbook = openpyxl.load_workbook('CUX_资产面积分配查询___021121.xlsx')
    active_sheet = workbook.active
    agents = list(active_sheet.columns)[column_index_from_string(AGENT_FIELD)-1]
    # 获取所有机构名称的前两个字作为分表标签：'六盘', '安顺', '毕节', '贵州', '贵阳'……
    agent_tags = set([agent.value[:2] for agent in agents[3:]])
    agent_workbooks = dict()

    # 根据机构名称建立分类Excel表
    for tag in agent_tags:
        agent_workbooks[tag] = openpyxl.Workbook()
        agent_workbooks[tag].active.append([title.value for title in active_sheet[3]])

    for row in active_sheet.iter_rows():
        area = row[column_index_from_string(AREA_FIELD)-1].value
        arch = row[column_index_from_string(ARCH_FIELD)-1].value
        agent = row[column_index_from_string(AGENT_FIELD)-1].value

        if type(area) != int and type(area) != float:
            continue

        if type(arch) != int and type(arch) != float:
            continue

        if (10 < area < 1000) and (10 < arch < 1000):
            continue

        agent_workbooks[agent[:2]].active.append([cell.value for cell in row])

    for tag, wb in agent_workbooks.items():
        wb.save('{}.xlsx'.format(tag))
        print('{}.xlsx saved'.format(tag))


if __name__ == '__main__':
    main()
