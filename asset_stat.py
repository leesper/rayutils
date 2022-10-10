import openpyxl
from openpyxl.utils import column_index_from_string as cifs

ASSET_NUM_COL = 'B' # 资产编号
ASSET_COM_COL = 'M' # 所属公司
ASSET_RES_COL = 'V' # 盘点结果
ASSET_AVL_COL = 'W' # 是否生效

RESULT_FILE = 'CUX_新资产盘点结果_10_10.xlsx'

def main():
    # 读excel表中所有记录
    source = openpyxl.load_workbook('CUX_新资产盘点表_10_10.xlsx')
    data = source.active

    # 过滤，只取是否生效为Y的记录，包括资产编号、所属公司和盘点结果
    avail_assets = []
    for row in data.iter_rows():
        asset_number = row[cifs(ASSET_NUM_COL)-1].value
        asset_company = row[cifs(ASSET_COM_COL)-1].value
        asset_result = row[cifs(ASSET_RES_COL)-1].value
        asset_avail = row[cifs(ASSET_AVL_COL)-1].value
        if asset_avail == 'Y':
            avail_assets.append((asset_number, asset_company, asset_result))

    # 统计所属公司下每个资产编号的盘点结果，包括待盘点、待审批、已盘点
    company2asset = dict()
    asset2stat = dict()
    for item in avail_assets:
        asset_number = item[0].strip()
        asset_company = item[1].strip()
        asset_result = item[2].strip()
        if not asset_company in company2asset.keys():
            company2asset[asset_company] = set()
        company2asset[asset_company].add(asset_number)

        if not asset_number in asset2stat.keys():
            asset2stat[asset_number] = dict()
            asset2stat[asset_number]['待审批'] = 0
            asset2stat[asset_number]['待盘点'] = 0
            asset2stat[asset_number]['已盘点'] = 0
        asset2stat[asset_number][asset_result] = asset2stat[asset_number].get(asset_result, 0) + 1

    # 待审批不为0的记录，全部算作待审批
    # 已盘点不为0的记录，全部算作已盘点
    # 其他的所有都是待盘点
    for stat in asset2stat.values():
        if stat['待审批'] > 0:
            stat['待审批'] += (stat['待盘点'] + stat['已盘点'])
            stat['待盘点'] = 0
            stat['已盘点'] = 0
        elif stat['已盘点'] > 0:
            stat['已盘点'] += (stat['待审批'] + stat['待盘点'])
            stat['待审批'] = 0
            stat['待盘点'] = 0
        else:
            stat['待盘点'] += (stat['待审批'] + stat['已盘点'])
            stat['待审批'] = 0
            stat['已盘点'] = 0
    
    # 统计每个单位的待盘点、待审批和已盘点数量，计算总数量和完成进度（已盘点/总数量百分比）写入结果表
    result_book = openpyxl.Workbook()
    result_active = result_book.active
    result_active.append(['单位', '总数量', '待盘点数量', '待审批数量', '已审批数量', '完成进度'])
    for company in company2asset.keys():
        total = 0
        to_check = 0
        to_appro = 0
        checked = 0
        for asset_number in company2asset[company]:
            stat = asset2stat[asset_number]
            to_check += stat['待盘点']
            to_appro += stat['待审批']
            checked += stat['已盘点']

        total += (to_check + to_appro + checked)
        progress = 1.0 * checked / total
        result_active.append([company, total, to_check, to_appro, checked, '{:.2%}'.format(progress)])
        # print('单位：{}\t总数量：{}\t待盘点数量：{}\t待审批数量：{}\t已审批数量：{}\t完成进度：{:.2%}\t'.format(company, total, to_check, to_appro, checked, progress))
    
    result_book.save(RESULT_FILE)
    print('{} saved'.format(RESULT_FILE))



if __name__ == '__main__':
    main()